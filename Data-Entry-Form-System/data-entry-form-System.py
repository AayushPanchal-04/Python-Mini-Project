from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook

# ===================== EXCEL FILE SETUP =====================

try:
    wb = openpyxl.load_workbook("data.xlsx")
    sheet = wb.active
except FileNotFoundError:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "User Data"

    sheet['A1'] = "Name"
    sheet['B1'] = "Age"
    sheet['C1'] = "Gender"
    sheet['D1'] = "Email"
    sheet['E1'] = "Contact"

    wb.save("data.xlsx")

# ===================== TKINTER WINDOW =====================

root = Tk()
root.title("User Data Entry System")
root.geometry("540x560")
root.resizable(False, False)
root.configure(bg="#f4f6f9")

# ===================== FUNCTIONS =====================

def clear_fields():
    name_entry.delete(0, END)
    age_entry.delete(0, END)
    gender_entry.delete(0, END)
    email_entry.delete(0, END)
    contact_entry.delete(0, END)
    name_entry.focus()

def submit_data():
    name = name_entry.get()
    age = age_entry.get()
    gender = gender_entry.get()
    email = email_entry.get()
    contact = contact_entry.get()

    if not all([name, age, gender, email, contact]):
        messagebox.showerror("Error", "All fields are required!")
        return

    row = sheet.max_row + 1

    sheet.cell(row=row, column=1).value = name
    sheet.cell(row=row, column=2).value = age
    sheet.cell(row=row, column=3).value = gender
    sheet.cell(row=row, column=4).value = email
    sheet.cell(row=row, column=5).value = contact

    wb.save("data.xlsx")
    messagebox.showinfo("Success", "Data saved successfully!")
    clear_fields()

# ===================== UI DESIGN =====================

# ---------- Header ----------
header = Frame(root, bg="#0b3c5d", height=65)
header.pack(fill=X)

Label(
    header,
    text="User Data Entry System",
    bg="#0b3c5d",
    fg="white",
    font=("Segoe UI", 18, "bold")
).pack(side=LEFT, padx=20, pady=15)

# ---------- Subtitle ----------
Label(
    root,
    text="Please fill in the information below",
    bg="#f4f6f9",
    fg="#555555",
    font=("Segoe UI", 11)
).pack(pady=(15, 5))

# ---------- Card ----------
card = Frame(
    root,
    bg="white",
    padx=40,
    pady=30,
    highlightthickness=1,
    highlightbackground="#d0d7de"
)
card.pack(pady=10)

label_font = ("Segoe UI", 10)
entry_font = ("Segoe UI", 11)

def make_label(text, row):
    Label(
        card,
        text=text,
        bg="white",
        fg="#222222",
        font=label_font
    ).grid(row=row, column=0, sticky=W, pady=12)

def make_entry(row):
    e = Entry(
        card,
        width=32,
        font=entry_font,
        relief=FLAT,
        bg="#f8f9fb",
        highlightthickness=1,
        highlightbackground="#cfd6dd",
        highlightcolor="#0b3c5d"
    )
    e.grid(row=row, column=1, pady=12, padx=(20, 0))
    return e

make_label("Full Name", 0)
name_entry = make_entry(0)

make_label("Age", 1)
age_entry = make_entry(1)

make_label("Gender", 2)
gender_entry = make_entry(2)

make_label("Email Address", 3)
email_entry = make_entry(3)

make_label("Contact Number", 4)
contact_entry = make_entry(4)

# ---------- Button ----------
Button(
    root,
    text="SAVE DATA",
    command=submit_data,
    font=("Segoe UI", 11, "bold"),
    bg="#0b3c5d",
    fg="white",
    activebackground="#082f48",
    activeforeground="white",
    relief=FLAT,
    width=20,
    cursor="hand2"
).pack(pady=25)

name_entry.focus()

# ===================== START GUI =====================
root.mainloop()
