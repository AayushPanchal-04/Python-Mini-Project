from tkinter import *
from tkinter import messagebox
import random, os
from openpyxl import Workbook, load_workbook

# ================= Window =================
root = Tk()
root.title("Bill Management System")
root.geometry("980x540")
root.resizable(False, False)
root.config(bg="#e6e6e6")

# ================= Prices =================
PRICE = {
    "Dosa": 60, "Cookies": 30, "Tea": 7,
    "Coffee": 100, "Juice": 20,
    "Pancakes": 15, "Eggs": 7
}

# ================= Variables =================
name = StringVar()
mobile = StringVar()
bill_no = StringVar(value=str(random.randint(1000, 9999)))
gst = DoubleVar(value=5)
discount = DoubleVar(value=0)
total = StringVar()

items = {k: IntVar() for k in PRICE}

# ================= Excel =================
FILE = "bills.xlsx"
if not os.path.exists(FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Bill No", "Name", "Mobile", "Subtotal", "GST", "Discount", "Net Total"])
    wb.save(FILE)

# ================= Functions =================
def calculate():
    subtotal = sum(items[i].get() * PRICE[i] for i in PRICE)
    gst_amt = subtotal * gst.get() / 100
    disc_amt = subtotal * discount.get() / 100
    net = subtotal + gst_amt - disc_amt

    total.set(f"Rs. {net:.2f}")
    generate_bill(subtotal, gst_amt, disc_amt, net)

def generate_bill(sub, gst_amt, disc_amt, net):
    receipt.delete("1.0", END)
    receipt.insert(END, "        BILL RECEIPT\n")
    receipt.insert(END, "-"*38 + "\n")
    receipt.insert(END, f"Bill No : {bill_no.get()}\n")
    receipt.insert(END, f"Name    : {name.get()}\n")
    receipt.insert(END, f"Mobile  : {mobile.get()}\n")
    receipt.insert(END, "-"*38 + "\n")

    for i in PRICE:
        if items[i].get() > 0:
            receipt.insert(
                END, f"{i:<12} x {items[i].get():<3} Rs.{items[i].get()*PRICE[i]}\n"
            )

    receipt.insert(END, "-"*38 + "\n")
    receipt.insert(END, f"Subtotal : Rs. {sub:.2f}\n")
    receipt.insert(END, f"GST ({gst.get()}%) : Rs. {gst_amt:.2f}\n")
    receipt.insert(END, f"Discount ({discount.get()}%) : Rs. {disc_amt:.2f}\n")
    receipt.insert(END, "-"*38 + "\n")
    receipt.insert(END, f"NET TOTAL : Rs. {net:.2f}\n")

def save_excel():
    if name.get()=="" or mobile.get()=="":
        messagebox.showerror("Error", "Enter customer details")
        return

    wb = load_workbook(FILE)
    ws = wb.active
    ws.append([
        bill_no.get(), name.get(), mobile.get(),
        "", gst.get(), discount.get(), total.get()
    ])
    wb.save(FILE)
    messagebox.showinfo("Saved", "Bill saved to Excel")

def print_bill():
    text = receipt.get("1.0", END)
    with open("print_bill.txt", "w") as f:
        f.write(text)
    os.startfile("print_bill.txt", "print")

def reset():
    name.set("")
    mobile.set("")
    gst.set(5)
    discount.set(0)
    total.set("")
    bill_no.set(str(random.randint(1000, 9999)))
    for i in items.values():
        i.set(0)
    receipt.delete("1.0", END)

# ================= Header =================
Label(
    root, text="BILL MANAGEMENT SYSTEM",
    bg="black", fg="white",
    font=("Arial", 22, "bold"),
    pady=8
).pack(fill=X)

# ================= Customer =================
cust = LabelFrame(root, text="Customer Details", bg="#f5f5f5",
                  font=("Arial", 11, "bold"), bd=3, relief=RIDGE)
cust.place(x=10, y=60, width=960, height=90)

Label(cust, text="Name", bg="#f5f5f5").grid(row=0, column=0, padx=15)
Entry(cust, textvariable=name, width=22).grid(row=0, column=1)

Label(cust, text="Mobile", bg="#f5f5f5").grid(row=0, column=2, padx=15)
Entry(cust, textvariable=mobile, width=22).grid(row=0, column=3)

Label(cust, text="Bill No", bg="#f5f5f5").grid(row=0, column=4, padx=15)
Entry(cust, textvariable=bill_no, width=15, state="readonly").grid(row=0, column=5)

Label(cust, text="GST %", bg="#f5f5f5").grid(row=1, column=0)
Entry(cust, textvariable=gst, width=10).grid(row=1, column=1)

Label(cust, text="Discount %", bg="#f5f5f5").grid(row=1, column=2)
Entry(cust, textvariable=discount, width=10).grid(row=1, column=3)

# ================= Menu =================
menu = LabelFrame(root, text="Menu", bg="#d1fae5",
                  font=("Arial", 12, "bold"), bd=3, relief=RIDGE)
menu.place(x=10, y=160, width=260, height=360)

Label(menu, text="""
Dosa........Rs.60
Cookies.....Rs.30
Tea.........Rs.7
Coffee......Rs.100
Juice.......Rs.20
Pancakes....Rs.15
Eggs........Rs.7
""", bg="#d1fae5", justify=LEFT, font=("Arial", 11)).pack(pady=10)

# ================= Items =================
entry = LabelFrame(root, text="Order Items", bg="#ffffff",
                   font=("Arial", 12, "bold"), bd=3, relief=RIDGE)
entry.place(x=280, y=160, width=300, height=360)

for i, item in enumerate(PRICE):
    Label(entry, text=item, bg="white", font=("Arial", 11, "bold")).grid(row=i, column=0, padx=20, pady=8, sticky="w")
    Entry(entry, textvariable=items[item], width=8, justify=CENTER).grid(row=i, column=1)

Button(entry, text="CALCULATE TOTAL", bg="#2563eb", fg="white",
       font=("Arial", 10, "bold"), width=18, command=calculate).grid(row=8, column=0, pady=15)

Button(entry, text="RESET", bg="#6b7280", fg="white",
       font=("Arial", 10, "bold"), width=18, command=reset).grid(row=8, column=1)

# ================= Bill =================
bill = LabelFrame(root, text="Bill Receipt", bg="#fff7ed",
                  font=("Arial", 12, "bold"), bd=3, relief=RIDGE)
bill.place(x=600, y=160, width=360, height=360)

receipt = Text(bill, width=42, height=15,
               font=("Consolas", 10), bg="white", relief=SUNKEN)
receipt.pack(pady=8)

Button(bill, text="SAVE BILL", bg="#16a34a", fg="white",
       font=("Arial", 10, "bold"), command=save_excel).pack(fill=X, padx=10, pady=4)

Button(bill, text="PRINT BILL", bg="#000000", fg="white",
       font=("Arial", 10, "bold"), command=print_bill).pack(fill=X, padx=10)

root.mainloop()
